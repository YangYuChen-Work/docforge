import json
from datetime import date, datetime
from types import SimpleNamespace

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.config import settings
from app.db.models import DocumentChapter, DocumentTemplate, GeneratedDocument
from app.db.session import Base
from app.domain import exports as domain_exports
from app.services.export_fonts import get_export_font_config
from app.services.xlsx_exporter import export_tables_to_xlsx


def _chapter(title, content_json):
    return SimpleNamespace(
        id=f"ch-{title}",
        title=title,
        content_json=content_json,
        missing_information_json="[]",
        conflict_json="[]",
    )


def _table(title, headers, rows):
    return {
        "title": title,
        "headers": headers,
        "rows": rows,
    }


def _load(path):
    return load_workbook(path)


def test_exporter_builds_overview_directory_and_table_sheets(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))

    chapters = [
        _chapter(
            "重复名称",
            json.dumps(
                {"tables": [_table("表A", ["序号", "名称"], [[1, "Alpha"]])]},
                ensure_ascii=False,
            ),
        ),
        _chapter(
            "重复名称",
            json.dumps(
                {"tables": [_table("表B", ["数值", "布尔", "日期", "空值", "嵌套"], [[3.5, True, "2026-08-09", None, {"x": 1}]])]},
                ensure_ascii=False,
            ),
        ),
        _chapter(
            "非法/名称:*?[]超长超长超长超长",
            json.dumps(
                {"tables": [_table("表C", ["序号"], [[2]])]},
                ensure_ascii=False,
            ),
        ),
        _chapter(
            "含有、逗号'章节",
            json.dumps(
                {"tables": [_table("表D", ["名称"], [["X"]])]},
                ensure_ascii=False,
            ),
        ),
        _chapter("坏JSON", "{not json"),
        _chapter("无表章", json.dumps({"paragraphs": []}, ensure_ascii=False)),
    ]

    out = export_tables_to_xlsx(
        "doc-1",
        chapters,
        document_meta={
            "title": "演示文档",
            "project_id": "P001",
            "template_name": "模板A",
            "status": "completed",
            "missing_items": ["待补充A", "待补充B"],
            "conflicts": ["冲突A"],
        },
    )

    wb = _load(out)
    assert wb.sheetnames[:2] == ["项目概览", "文档目录"]
    assert wb.sheetnames[2:] == [
        "重复名称",
        "重复名称-2",
        "非法名称超长超长超长超长",
        "含有、逗号'章节",
    ]

    overview = wb["项目概览"]
    assert overview["A1"].value == "导出概览"
    assert overview["A3"].value == "文档标题"
    assert overview["B3"].value == "演示文档"
    assert overview["A4"].value == "项目编号"
    assert overview["B4"].value == "P001"
    assert overview["A6"].value == "状态"
    assert overview["B6"].value == "completed"
    assert overview["A7"].value == "待补充项"
    assert "待补充A" in overview["B7"].value
    assert overview["A8"].value == "冲突项"
    assert "冲突A" in overview["B8"].value

    directory = wb["文档目录"]
    assert directory["A1"].value == "章节"
    assert directory["B1"].value == "表格数量"
    assert directory["C1"].value == "Sheet 名称"
    assert directory["D1"].value == "打开"

    rows = list(directory.iter_rows(min_row=2, max_row=5, values_only=True))
    assert rows[0][:3] == ("重复名称", 1, "重复名称")
    assert rows[0][3].startswith("=HYPERLINK(")
    assert rows[1][:3] == ("重复名称", 1, "重复名称-2")
    assert rows[1][3].startswith("=HYPERLINK(")
    assert rows[2][:3] == ("非法/名称:*?[]超长超长超长超长", 1, "非法名称超长超长超长超长")
    assert rows[2][3].startswith("=HYPERLINK(")
    assert rows[3][:3] == ("含有、逗号'章节", 1, "含有、逗号'章节")
    assert rows[3][3] == '=HYPERLINK("#\'含有、逗号\'\'章节\'!A1","打开")'

    sheet = wb["重复名称"]
    assert sheet.freeze_panes == "A4"
    assert sheet.auto_filter.ref == "A3:B4"
    assert sheet["A1"].value == "表A"
    assert sheet["A2"].value == "来源章节：重复名称"
    assert sheet["A3"].value == "序号"
    assert sheet["B3"].value == "名称"
    assert sheet["A4"].value == 1
    assert sheet["B4"].value == "Alpha"
    assert sheet["A3"].fill.fgColor.rgb in ("FFD9EAF7", "00D9EAF7")
    assert sheet["A4"].alignment.wrap_text is True
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1

    second = wb["重复名称-2"]
    assert second["A4"].value == 3.5
    assert second["B4"].value is True
    assert second["C4"].is_date
    normalized = second["C4"].value
    if hasattr(normalized, "date"):
        normalized = normalized.date()
    assert normalized == date(2026, 8, 9)
    assert second["D4"].value is None
    assert second["E4"].value == '{"x": 1}'

    long_name = wb["非法名称超长超长超长超长"]
    assert len(long_name.title) <= 31
    assert "/" not in long_name.title
    assert ":" not in long_name.title
    assert "*" not in long_name.title
    assert "?" not in long_name.title
    assert "[" not in long_name.title
    assert "]" not in long_name.title


def test_exporter_uses_case_insensitive_unique_sheet_names_in_directory_links(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))

    chapters = [
        _chapter(
            "Report",
            json.dumps({"tables": [_table("Upper", ["Value"], [[1]])]}, ensure_ascii=False),
        ),
        _chapter(
            "report",
            json.dumps({"tables": [_table("Lower", ["Value"], [[2]])]}, ensure_ascii=False),
        ),
    ]

    out = export_tables_to_xlsx("doc-case-collision", chapters)
    wb = _load(out)

    actual_sheet_names = wb.sheetnames[2:]
    assert actual_sheet_names == ["Report", "report-2"]
    assert len({name.casefold() for name in actual_sheet_names}) == len(actual_sheet_names)

    directory = wb["文档目录"]
    assert directory["C2"].value == actual_sheet_names[0]
    assert directory["D2"].value == '=HYPERLINK("#\'Report\'!A1","打开")'
    assert directory["C3"].value == actual_sheet_names[1]
    assert directory["D3"].value == '=HYPERLINK("#\'report-2\'!A1","打开")'


def test_exporter_escapes_double_quotes_in_directory_links(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))

    out = export_tables_to_xlsx(
        "doc-quoted-sheet",
        [
            _chapter(
                '含"双引号"章节',
                json.dumps(
                    {"tables": [_table("表A", ["值"], [[1]])]},
                    ensure_ascii=False,
                ),
            )
        ],
    )

    wb = _load(out)
    directory = wb["文档目录"]
    assert wb.sheetnames[2] == '含"双引号"章节'
    assert directory["D2"].value == '=HYPERLINK("#\'含""双引号""章节\'!A1","打开")'


def test_create_export_passes_xlsx_metadata_and_issue_summary(tmp_path, monkeypatch):
    engine = create_engine(
        f"sqlite:///{tmp_path / 'export-meta.db'}",
        connect_args={"check_same_thread": False},
    )
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine, autocommit=False, autoflush=False)
    db = Session()

    tpl = DocumentTemplate(
        id="TPL1",
        name="模板A",
        phase="phase1",
        category="cat1",
        source_path=None,
        export_format="xlsx",
    )
    doc = GeneratedDocument(
        id="DOC1",
        project_id="P001",
        generation_task_id="TASK1",
        template_id="TPL1",
        title="演示文档",
        status="confirmed",
    )
    chapters = [
        DocumentChapter(
            id="CH1",
            document_id="DOC1",
            title="章节1",
            order_index=1,
            status="confirmed",
            missing_information_json="{bad json",
            conflict_json="{bad json",
        ),
        DocumentChapter(
            id="CH2",
            document_id="DOC1",
            title="章节2",
            order_index=2,
            status="confirmed",
            missing_information_json=json.dumps(
                [
                    "缺失1",
                    {"description": ["缺失2", {"source": "S1"}]},
                    {"title": {"label": "缺失3"}},
                    {"name": 42},
                    "缺失5",
                    "缺失6",
                ],
                ensure_ascii=False,
            ),
            conflict_json=json.dumps(
                [
                    {"description": "冲突1"},
                    {"description": ["冲突2"]},
                    {"title": {"label": "冲突3"}},
                    {"name": 42},
                    {"description": "冲突5"},
                    {"description": "冲突6"},
                ],
                ensure_ascii=False,
            ),
        ),
    ]
    db.add_all([tpl, doc, *chapters])
    db.commit()

    captured = {}

    monkeypatch.setattr(
        domain_exports,
        "validate_document",
        lambda chapters, template_path: {
            "passed": True,
            "warnings": [],
            "errors": [],
            "can_export": True,
            "has_missing_info": True,
        },
    )

    def fake_export_tables_to_xlsx(doc_id, chapters, document_meta=None):
        captured["doc_id"] = doc_id
        captured["chapters"] = chapters
        captured["document_meta"] = document_meta
        fake_path = tmp_path / "fake.xlsx"
        fake_path.write_bytes(b"fake")
        return str(fake_path)

    monkeypatch.setattr(
        "app.services.xlsx_exporter.export_tables_to_xlsx",
        fake_export_tables_to_xlsx,
    )

    export = domain_exports.create_export(db, "DOC1", "xlsx")

    assert export.status == "completed"
    assert export.output_path == str(tmp_path / "fake.xlsx")
    assert captured["doc_id"] == "DOC1"
    assert [ch.id for ch in captured["chapters"]] == ["CH1", "CH2"]
    meta = captured["document_meta"]
    assert meta["title"] == "演示文档"
    assert meta["project_id"] == "P001"
    assert meta["template_name"] == "模板A"
    assert meta["status"] == "confirmed"
    assert meta["missing_items"] == [
        "缺失1",
        '["缺失2", {"source": "S1"}]',
        '{"label": "缺失3"}',
        "42",
        "缺失5",
    ]
    assert meta["conflicts"] == [
        "冲突1",
        '["冲突2"]',
        '{"label": "冲突3"}',
        "42",
        "冲突5",
    ]
    db.close()


def test_exporter_normalizes_issue_metadata_before_overview_truncation(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))
    issue_items = [
        {"description": ["列表描述"]},
        {"title": {"label": "字典标题"}},
        {"name": 42},
        {"description": "x" * 70},
    ]

    out = export_tables_to_xlsx(
        "doc-metadata-values",
        [],
        document_meta={
            "title": {"label": "演示文档"},
            "project_id": ["P001"],
            "template_name": 7,
            "status": False,
            "missing_items": issue_items,
            "conflicts": issue_items,
        },
    )

    overview = _load(out)["项目概览"]
    assert overview["B3"].value == '{"label": "演示文档"}'
    assert overview["B4"].value == '["P001"]'
    assert overview["B5"].value == "7"
    assert overview["B6"].value == "False"
    assert overview["B7"].value.startswith(
        '["列表描述"]；{"label": "字典标题"}；42；'
    )
    assert overview["B7"].value.endswith("...")
    assert overview["B8"].value == overview["B7"].value


def test_exporter_falls_back_when_no_valid_tables_exist(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))

    chapters = [
        _chapter("坏JSON", "{not json"),
        _chapter("空章节", json.dumps({"content": []}, ensure_ascii=False)),
    ]

    out = export_tables_to_xlsx("doc-2", chapters, document_meta=None)
    wb = _load(out)

    assert wb.sheetnames == ["项目概览", "文档目录", "无表格数据"]
    fallback = wb["无表格数据"]
    assert fallback["A1"].value == "本文档暂无可导出的表格数据"
    assert fallback["A2"].value == "请检查章节内容是否包含 tables 节点，或等待章节重新生成。"


def test_exporter_ignores_legacy_tables_with_null_optional_fields(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))
    content = {
        "tables": [
            {"title": "行为空", "headers": ["名称"], "rows": None},
            {"title": "表头为空", "headers": None, "rows": [["产品"]]},
        ]
    }

    out = export_tables_to_xlsx(
        "doc-null-legacy-fields",
        [_chapter("异常表格", json.dumps(content, ensure_ascii=False))],
    )

    wb = _load(out)
    assert wb.sheetnames == ["项目概览", "文档目录", "无表格数据"]


def test_exporter_safely_writes_malformed_legacy_table_metadata(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))
    content = {
        "tables": [
            {
                "title": ["错误标题"],
                "caption": "可用标题",
                "headers": ["名称"],
                "rows": [["产品A"]],
            },
            {
                "title": {"label": "字典标题"},
                "headers": ["名称"],
                "rows": [["产品B"]],
            },
            {
                "caption": ["列表标题", 2],
                "headers": ["名称"],
                "rows": [["产品C"]],
            },
            {
                "title": 42,
                "headers": ["名称"],
                "rows": [["产品D"]],
            },
            {
                "title": [],
                "caption": {},
                "headers": ["名称"],
                "rows": [["产品E"]],
            },
        ]
    }

    out = export_tables_to_xlsx(
        "doc-malformed-legacy-metadata",
        [_chapter("异常表格", json.dumps(content, ensure_ascii=False))],
    )

    wb = _load(out)
    assert wb.sheetnames == [
        "项目概览",
        "文档目录",
        "异常表格-1",
        "异常表格-2",
        "异常表格-3",
        "异常表格-4",
        "异常表格-5",
    ]
    assert wb["异常表格-1"]["A1"].value == "可用标题"
    assert wb["异常表格-2"]["A1"].value == '{"label": "字典标题"}'
    assert wb["异常表格-3"]["A1"].value == '["列表标题", 2]'
    assert wb["异常表格-4"]["A1"].value == "42"
    assert wb["异常表格-5"]["A1"].value == "异常表格 表格"


def test_exporter_ignores_prosemirror_tables_with_null_content(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))
    content = {
        "type": "doc",
        "content": [
            {"type": "table", "content": None},
            {
                "type": "table",
                "content": [
                    {"type": "tableRow", "content": None},
                    {
                        "type": "tableRow",
                        "content": [
                            {"type": "tableHeader", "content": None},
                        ],
                    },
                ],
            },
        ],
    }

    out = export_tables_to_xlsx(
        "doc-null-prosemirror-content",
        [_chapter("异常表格", json.dumps(content, ensure_ascii=False))],
    )

    wb = _load(out)
    assert wb.sheetnames == ["项目概览", "文档目录", "无表格数据"]


def test_xlsx_cells_use_explicit_cjk_font(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))
    content = json.dumps(
        {"tables": [_table("中文表格", ["名称"], [["产品"]])]},
        ensure_ascii=False,
    )

    out = export_tables_to_xlsx("doc-fonts", [_chapter("产品概述", content)])
    wb = _load(out)
    fonts = get_export_font_config()

    sheet = wb["产品概述"]
    assert sheet["A1"].font.name == fonts.cjk
    assert sheet["A2"].font.name == fonts.cjk
    assert sheet["A3"].font.name == fonts.cjk
    assert sheet["A4"].font.name == fonts.cjk


def test_exporter_reads_production_prosemirror_tables_and_preserves_types(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "storage_root", str(tmp_path))
    content = {
        "type": "doc",
        "content": [
            {
                "type": "paragraph",
                "content": [{"type": "text", "text": "核心性能参数表"}],
            },
            {
                "type": "table",
                "content": [
                    {
                        "type": "tableRow",
                        "content": [
                            {
                                "type": "tableHeader",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "数量"}]}],
                            },
                            {
                                "type": "tableHeader",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "启用"}]}],
                            },
                            {
                                "type": "tableHeader",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "更新时间"}]}],
                            },
                            {
                                "type": "tableHeader",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "说明"}]}],
                            },
                        ],
                    },
                    {
                        "type": "tableRow",
                        "content": [
                            {
                                "type": "tableCell",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "80"}]}],
                            },
                            {
                                "type": "tableCell",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "true"}]}],
                            },
                            {
                                "type": "tableCell",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "2026-08-09T20:15:00+08:00"}]}],
                            },
                            {
                                "type": "tableCell",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "主臂系统"}]}],
                            },
                        ],
                    },
                    {
                        "type": "tableRow",
                        "content": [
                            {
                                "type": "tableCell",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "81"}]}],
                            },
                            {
                                "type": "tableCell",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "false"}]}],
                            },
                            {
                                "type": "tableCell",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "2026-08-10T01:02:03Z"}]}],
                            },
                            {
                                "type": "tableCell",
                                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "副臂系统"}]}],
                            },
                        ],
                    },
                ],
            },
        ],
    }

    out = export_tables_to_xlsx(
        "doc-prosemirror",
        [_chapter("产品参数", json.dumps(content, ensure_ascii=False))],
    )

    wb = _load(out)
    assert wb.sheetnames == ["项目概览", "文档目录", "产品参数"]
    sheet = wb["产品参数"]
    assert sheet["A1"].value == "核心性能参数表"
    assert sheet["A2"].value == "来源章节：产品参数"
    assert [sheet.cell(row=3, column=column).value for column in range(1, 5)] == [
        "数量",
        "启用",
        "更新时间",
        "说明",
    ]
    assert sheet["A4"].value == 80
    assert sheet["A4"].data_type == "n"
    assert sheet["B4"].value is True
    assert sheet["B4"].data_type == "b"
    assert sheet["C4"].value == datetime(2026, 8, 9, 12, 15)
    assert sheet["C4"].value.tzinfo is None
    assert sheet["C4"].is_date
    assert sheet["D4"].value == "主臂系统"
    assert sheet["C5"].value == datetime(2026, 8, 10, 1, 2, 3)
    assert sheet["C5"].value.tzinfo is None
    assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
