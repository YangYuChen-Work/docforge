import json
from collections.abc import Iterable
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import get_storage_path


PROJECT_BLUE_FILL = PatternFill("solid", fgColor="1F4E78")
PALE_BLUE_FILL = PatternFill("solid", fgColor="D9EAF7")
NEUTRAL_FILL = PatternFill("solid", fgColor="F8FAFC")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
PROJECT_BLUE_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(color="1F1F1F", bold=True)
BODY_FONT = Font(color="1F1F1F")
NEUTRAL_BORDER = Border(
    left=Side(style="thin", color="B7C0CC"),
    right=Side(style="thin", color="B7C0CC"),
    top=Side(style="thin", color="B7C0CC"),
    bottom=Side(style="thin", color="B7C0CC"),
)
CENTER_WRAP = Alignment(vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LINK_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def export_tables_to_xlsx(doc_id: str, chapters: list, document_meta: dict | None = None) -> str:
    out_dir = get_storage_path("generated") / doc_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "tables.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    document_meta = document_meta or {}
    export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    overview = wb.create_sheet("项目概览")
    _build_overview_sheet(overview, document_meta, export_time)

    directory = wb.create_sheet("文档目录")
    _build_directory_header(directory)

    used_sheet_names: set[str] = set(overview.title for overview in wb.worksheets)
    directory_rows: list[dict[str, object]] = []
    has_table = False

    for chapter in chapters:
        tables = _load_tables(getattr(chapter, "content_json", None))
        if not tables:
            continue

        chapter_sheet_names: list[str] = []
        for table_index, table in enumerate(tables, start=1):
            base_name = _sanitize_sheet_name(getattr(chapter, "title", "") or "Sheet")
            if len(tables) > 1:
                base_name = _sanitize_sheet_name(f"{base_name}-{table_index}")
            sheet_name = _unique_sheet_name(base_name, used_sheet_names)
            chapter_sheet_names.append(sheet_name)

            ws = wb.create_sheet(title=sheet_name)
            _write_table_sheet(ws, chapter, table)
            has_table = True

        directory_rows.append(
            {
                "chapter_title": getattr(chapter, "title", "") or "未命名章节",
                "table_count": len(chapter_sheet_names),
                "sheet_names": chapter_sheet_names,
            }
        )

    for row_index, entry in enumerate(directory_rows, start=2):
        chapter_title = entry["chapter_title"]
        table_count = entry["table_count"]
        sheet_names = entry["sheet_names"]
        first_sheet_name = sheet_names[0] if sheet_names else ""
        directory.cell(row=row_index, column=1, value=chapter_title)
        directory.cell(row=row_index, column=2, value=table_count)
        directory.cell(row=row_index, column=3, value="、".join(sheet_names))
        open_cell = directory.cell(row=row_index, column=4)
        if first_sheet_name:
            open_cell.value = f'=HYPERLINK("{_excel_sheet_target(first_sheet_name)}","打开")'
        for col in range(1, 5):
            cell = directory.cell(row=row_index, column=col)
            cell.font = BODY_FONT
            cell.border = NEUTRAL_BORDER
            cell.alignment = LEFT_WRAP if col != 4 else LINK_ALIGNMENT
            if col == 4 and first_sheet_name:
                cell.font = Font(color="0563C1", underline="single")

    if not has_table:
        ws = wb.create_sheet(title="无表格数据")
        ws["A1"] = "本文档暂无可导出的表格数据"
        ws["A2"] = "请检查章节内容是否包含 tables 节点，或等待章节重新生成。"
        for cell in ("A1", "A2"):
            ws[cell].font = BODY_FONT
            ws[cell].alignment = LEFT_WRAP
        ws.column_dimensions["A"].width = 72
    wb.save(str(out_path))
    return str(out_path)


def _build_overview_sheet(ws, document_meta: dict, export_time: str) -> None:
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "导出概览"
    title.fill = PROJECT_BLUE_FILL
    title.font = Font(color="FFFFFF", bold=True, size=14)
    title.alignment = CENTER_WRAP
    title.border = NEUTRAL_BORDER

    rows = [
        ("文档标题", document_meta.get("title", "")),
        ("项目编号", document_meta.get("project_id", "")),
        ("模板名称", document_meta.get("template_name", "")),
        ("状态", document_meta.get("status", "")),
        ("待补充项", _summarize_items(document_meta.get("missing_items"))),
        ("冲突项", _summarize_items(document_meta.get("conflicts"))),
        ("导出时间", export_time),
    ]

    for row_index, (label, value) in enumerate(rows, start=3):
        label_cell = ws.cell(row=row_index, column=1, value=label)
        value_cell = ws.cell(row=row_index, column=2, value=value or "无")
        label_cell.fill = PALE_BLUE_FILL
        label_cell.font = SECTION_FONT
        label_cell.border = NEUTRAL_BORDER
        label_cell.alignment = LEFT_WRAP
        value_cell.fill = WARNING_FILL if label in ("待补充项", "冲突项") and value not in ("", "无") else NEUTRAL_FILL
        value_cell.font = BODY_FONT
        value_cell.border = NEUTRAL_BORDER
        value_cell.alignment = LEFT_WRAP

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 72
    ws.row_dimensions[1].height = 24


def _build_directory_header(ws) -> None:
    headers = ("章节", "表格数量", "Sheet 名称", "打开")
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = PROJECT_BLUE_FILL
        cell.font = PROJECT_BLUE_FONT
        cell.border = NEUTRAL_BORDER
        cell.alignment = CENTER_WRAP

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 12


def _write_table_sheet(ws, chapter, table: dict) -> None:
    headers = _normalize_row(table.get("headers", []))
    rows = [_normalize_row(row) for row in table.get("rows", [])]
    max_cols = max([len(headers)] + [len(row) for row in rows] if rows else [len(headers), 1])
    headers = _pad_row(headers, max_cols)
    rows = [_pad_row(row, max_cols) for row in rows]

    table_title = table.get("title") or table.get("caption") or f"{getattr(chapter, 'title', '章节')} 表格"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)

    title_cell = ws.cell(row=1, column=1, value=table_title)
    source_cell = ws.cell(row=2, column=1, value=f"来源章节：{getattr(chapter, 'title', '')}")
    title_cell.fill = PROJECT_BLUE_FILL
    title_cell.font = Font(color="FFFFFF", bold=True)
    title_cell.alignment = LEFT_WRAP
    title_cell.border = NEUTRAL_BORDER
    source_cell.fill = PALE_BLUE_FILL
    source_cell.font = BODY_FONT
    source_cell.alignment = LEFT_WRAP
    source_cell.border = NEUTRAL_BORDER

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_index, value=header)
        cell.fill = PALE_BLUE_FILL
        cell.font = SECTION_FONT
        cell.border = NEUTRAL_BORDER
        cell.alignment = CENTER_WRAP

    for row_index, row in enumerate(rows, start=4):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.font = BODY_FONT
            cell.border = NEUTRAL_BORDER
            cell.alignment = LEFT_WRAP

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(max_cols)}{max(3, len(rows) + 3)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:3"

    _set_column_widths(ws, max_cols, headers, rows)


def _load_tables(content_json: str | None) -> list[dict]:
    if not content_json:
        return []
    try:
        data = json.loads(content_json)
    except (TypeError, json.JSONDecodeError):
        return []
    if not isinstance(data, dict):
        return []

    tables = data.get("tables", [])
    if not isinstance(tables, list):
        return []
    return [table for table in tables if isinstance(table, dict) and table.get("headers")]


def _normalize_row(row: Iterable) -> list:
    if not isinstance(row, Iterable) or isinstance(row, (str, bytes, dict)):
        return [_normalize_cell_value(row)]
    return [_normalize_cell_value(value) for value in row]


def _normalize_cell_value(value):
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, str):
        parsed = _parse_iso_date(value)
        return parsed if parsed is not None else value
    if isinstance(value, (list, tuple, set, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _parse_iso_date(value: str):
    try:
        if len(value) == 10 and value[4] == "-" and value[7] == "-":
            return date.fromisoformat(value)
        if "T" in value or " " in value:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    return None


def _pad_row(row: list, size: int) -> list:
    return row + [None] * max(0, size - len(row))


def _sanitize_sheet_name(name: str) -> str:
    cleaned = "".join(ch for ch in name if ch not in "\\/:*?[]")
    cleaned = cleaned.strip() or "Sheet"
    return cleaned[:31]


def _excel_sheet_target(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"#'{escaped}'!A1"


def _unique_sheet_name(base_name: str, used_names: set[str]) -> str:
    candidate = base_name[:31] or "Sheet"
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate

    suffix_number = 2
    while True:
        suffix = f"-{suffix_number}"
        trimmed = base_name[: max(1, 31 - len(suffix))].rstrip("-")
        candidate = f"{trimmed}{suffix}"[:31]
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        suffix_number += 1


def _set_column_widths(ws, max_cols: int, headers: list, rows: list[list]) -> None:
    all_rows = [headers, *rows]
    for col_index in range(1, max_cols + 1):
        values = []
        for row in all_rows:
            if col_index - 1 < len(row):
                values.append(row[col_index - 1])
        width = max([len(str(v)) for v in values if v not in (None, "")] or [10])
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(width + 2, 10), 28)


def _summarize_items(value) -> str:
    if not value:
        return "无"
    if isinstance(value, str):
        items = [value]
    elif isinstance(value, Iterable):
        items = []
        for item in value:
            if isinstance(item, str):
                text = item.strip()
            elif isinstance(item, dict):
                text = (
                    item.get("description")
                    or item.get("title")
                    or item.get("name")
                    or json.dumps(item, ensure_ascii=False)
                )
            else:
                text = str(item)
            if text:
                items.append(text)
    else:
        items = [str(value)]

    compact = [item if len(item) <= 60 else item[:57] + "..." for item in items[:5]]
    if len(items) > 5:
        compact.append(f"等{len(items) - 5}项")
    return "；".join(compact) or "无"
