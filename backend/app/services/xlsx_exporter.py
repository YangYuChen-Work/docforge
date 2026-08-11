import json
import re
from collections.abc import Iterable
from datetime import date, datetime, timezone

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import get_storage_path
from app.services.export_fonts import get_export_font_config


EXPORT_FONTS = get_export_font_config()


def _cell_font(**kwargs) -> Font:
    return Font(name=EXPORT_FONTS.cjk, **kwargs)


PROJECT_BLUE_FILL = PatternFill("solid", fgColor="1F4E78")
PALE_BLUE_FILL = PatternFill("solid", fgColor="D9EAF7")
NEUTRAL_FILL = PatternFill("solid", fgColor="F8FAFC")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
PROJECT_BLUE_FONT = _cell_font(color="FFFFFF", bold=True)
SECTION_FONT = _cell_font(color="1F1F1F", bold=True)
BODY_FONT = _cell_font(color="1F1F1F")
NEUTRAL_BORDER = Border(
    left=Side(style="thin", color="B7C0CC"),
    right=Side(style="thin", color="B7C0CC"),
    top=Side(style="thin", color="B7C0CC"),
    bottom=Side(style="thin", color="B7C0CC"),
)
CENTER_WRAP = Alignment(vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LINK_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def export_tables_to_xlsx(
    doc_id: str,
    chapters: list,
    document_meta: dict | None = None,
    annotations: list | None = None,
) -> str:
    out_dir = get_storage_path("generated") / doc_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "tables.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    document_meta = document_meta or {}
    export_annotations = _active_export_annotations(annotations)
    annotations_by_chapter = _annotations_by_chapter(export_annotations)
    comment_locations: dict[str, str] = {}
    export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    overview = wb.create_sheet("项目概览")
    _build_overview_sheet(overview, document_meta, export_time)

    directory = wb.create_sheet("文档目录")
    _build_directory_header(directory)

    used_sheet_names: set[str] = set(overview.title for overview in wb.worksheets)
    if export_annotations:
        used_sheet_names.add("批注")
    directory_rows: list[dict[str, object]] = []
    has_table = False

    for chapter in chapters:
        tables = _load_tables(getattr(chapter, "content_json", None))
        if not tables:
            continue

        chapter_sheet_names: list[str] = []
        for table_index, table in enumerate(tables, start=1):
            base_name = _sanitize_sheet_name(getattr(chapter, "title", "") or "Sheet")
            if len(tables) > 1:
                base_name = _sanitize_sheet_name(f"{base_name}-{table_index}")
            sheet_name = _unique_sheet_name(base_name, used_sheet_names)

            ws = wb.create_sheet(title=sheet_name)
            actual_sheet_name = ws.title
            chapter_sheet_names.append(actual_sheet_name)
            _write_table_sheet(
                ws,
                chapter,
                table,
                annotations_by_chapter.get(str(getattr(chapter, "id", "")), []),
                comment_locations,
            )
            has_table = True

        directory_rows.append(
            {
                "chapter_title": getattr(chapter, "title", "") or "未命名章节",
                "table_count": len(chapter_sheet_names),
                "sheet_names": chapter_sheet_names,
            }
        )

    for row_index, entry in enumerate(directory_rows, start=2):
        chapter_title = entry["chapter_title"]
        table_count = entry["table_count"]
        sheet_names = entry["sheet_names"]
        first_sheet_name = sheet_names[0] if sheet_names else ""
        directory.cell(row=row_index, column=1, value=chapter_title)
        directory.cell(row=row_index, column=2, value=table_count)
        directory.cell(row=row_index, column=3, value="、".join(sheet_names))
        open_cell = directory.cell(row=row_index, column=4)
        if first_sheet_name:
            open_cell.value = f'=HYPERLINK("{_excel_sheet_target(first_sheet_name)}","打开")'
        for col in range(1, 5):
            cell = directory.cell(row=row_index, column=col)
            cell.font = BODY_FONT
            cell.border = NEUTRAL_BORDER
            cell.alignment = LEFT_WRAP if col != 4 else LINK_ALIGNMENT
            if col == 4 and first_sheet_name:
                cell.font = _cell_font(color="0563C1", underline="single")

    if not has_table:
        ws = wb.create_sheet(title="无表格数据")
        ws["A1"] = "本文档暂无可导出的表格数据"
        ws["A2"] = "请检查章节内容是否包含 tables 节点，或等待章节重新生成。"
        for cell in ("A1", "A2"):
            ws[cell].font = BODY_FONT
            ws[cell].alignment = LEFT_WRAP
        ws.column_dimensions["A"].width = 72
    if export_annotations:
        _build_comments_sheet(wb, export_annotations, comment_locations)
    wb.save(str(out_path))
    return str(out_path)


def _build_overview_sheet(ws, document_meta: dict, export_time: str) -> None:
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "导出概览"
    title.fill = PROJECT_BLUE_FILL
    title.font = _cell_font(color="FFFFFF", bold=True, size=14)
    title.alignment = CENTER_WRAP
    title.border = NEUTRAL_BORDER

    rows = [
        ("文档标题", _normalize_metadata_text(document_meta.get("title", ""))),
        ("项目编号", _normalize_metadata_text(document_meta.get("project_id", ""))),
        ("模板名称", _normalize_metadata_text(document_meta.get("template_name", ""))),
        ("状态", _normalize_metadata_text(document_meta.get("status", ""))),
        ("待补充项", _summarize_items(document_meta.get("missing_items"))),
        ("冲突项", _summarize_items(document_meta.get("conflicts"))),
        ("导出时间", export_time),
    ]

    for row_index, (label, value) in enumerate(rows, start=3):
        label_cell = ws.cell(row=row_index, column=1, value=label)
        value_cell = ws.cell(row=row_index, column=2, value=value or "无")
        label_cell.fill = PALE_BLUE_FILL
        label_cell.font = SECTION_FONT
        label_cell.border = NEUTRAL_BORDER
        label_cell.alignment = LEFT_WRAP
        value_cell.fill = WARNING_FILL if label in ("待补充项", "冲突项") and value not in ("", "无") else NEUTRAL_FILL
        value_cell.font = BODY_FONT
        value_cell.border = NEUTRAL_BORDER
        value_cell.alignment = LEFT_WRAP

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 72
    ws.row_dimensions[1].height = 24


def _build_directory_header(ws) -> None:
    headers = ("章节", "表格数量", "Sheet 名称", "打开")
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = PROJECT_BLUE_FILL
        cell.font = PROJECT_BLUE_FONT
        cell.border = NEUTRAL_BORDER
        cell.alignment = CENTER_WRAP

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 12


def _write_table_sheet(
    ws,
    chapter,
    table: dict,
    annotations: list[dict] | None = None,
    comment_locations: dict[str, str] | None = None,
) -> None:
    raw_headers = table.get("headers")
    raw_rows = table.get("rows")
    headers = _normalize_row(raw_headers) if isinstance(raw_headers, list) else []
    rows = [_normalize_row(row) for row in raw_rows] if isinstance(raw_rows, list) else []
    max_cols = max([len(headers)] + [len(row) for row in rows] if rows else [len(headers), 1])
    headers = _pad_row(headers, max_cols)
    rows = [_pad_row(row, max_cols) for row in rows]

    table_title = _legacy_table_title(table, chapter)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)

    title_cell = ws.cell(row=1, column=1, value=table_title)
    source_cell = ws.cell(row=2, column=1, value=f"来源章节：{getattr(chapter, 'title', '')}")
    title_cell.fill = PROJECT_BLUE_FILL
    title_cell.font = _cell_font(color="FFFFFF", bold=True)
    title_cell.alignment = LEFT_WRAP
    title_cell.border = NEUTRAL_BORDER
    source_cell.fill = PALE_BLUE_FILL
    source_cell.font = BODY_FONT
    source_cell.alignment = LEFT_WRAP
    source_cell.border = NEUTRAL_BORDER

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_index, value=header)
        cell.fill = PALE_BLUE_FILL
        cell.font = SECTION_FONT
        cell.border = NEUTRAL_BORDER
        cell.alignment = CENTER_WRAP

    for row_index, row in enumerate(rows, start=4):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.font = BODY_FONT
            cell.border = NEUTRAL_BORDER
            cell.alignment = LEFT_WRAP

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(max_cols)}{max(3, len(rows) + 3)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:3"

    _set_column_widths(ws, max_cols, headers, rows)
    _attach_cell_comments(
        ws,
        annotations or [],
        comment_locations if comment_locations is not None else {},
    )


def _attach_cell_comments(ws, annotations: list[dict], comment_locations: dict[str, str]) -> None:
    for annotation in annotations:
        annotation_id = str(annotation.get("id") or "")
        if annotation_id in comment_locations:
            continue
        target_text = annotation.get("target_text", "")
        if not target_text:
            continue
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if target_text not in str(cell.value or ""):
                    continue
                content = annotation.get("content") or "（未填写批示内容）"
                author = str(annotation.get("created_by") or "本地用户")
                cell.comment = Comment(content, author)
                comment_locations[annotation_id] = f"{ws.title}!{cell.coordinate}"
                break
            if annotation_id in comment_locations:
                break


def _build_comments_sheet(wb: Workbook, annotations: list[dict], comment_locations: dict[str, str]) -> None:
    ws = wb.create_sheet("批注")
    ws.merge_cells("A1:F1")
    ws["A1"] = "批注清单"
    ws["A1"].fill = PROJECT_BLUE_FILL
    ws["A1"].font = _cell_font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = CENTER_WRAP
    ws["A1"].border = NEUTRAL_BORDER

    headers = ("原文", "批注内容", "章节", "状态", "定位说明", "位置")
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=column, value=header)
        cell.fill = PALE_BLUE_FILL
        cell.font = SECTION_FONT
        cell.border = NEUTRAL_BORDER
        cell.alignment = CENTER_WRAP

    for row_index, annotation in enumerate(annotations, start=4):
        annotation_id = str(annotation.get("id") or "")
        values = (
            annotation.get("target_text") or "未提供原文",
            annotation.get("content") or "（未填写批示内容）",
            annotation.get("chapter_title") or annotation.get("chapter_id") or "未命名章节",
            annotation.get("status") or "pending",
            annotation.get("locator") or "未提供定位",
            comment_locations.get(annotation_id, "未在表格中定位"),
        )
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.font = BODY_FONT
            cell.border = NEUTRAL_BORDER
            cell.alignment = LEFT_WRAP

    widths = (28, 42, 24, 12, 24, 24)
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{max(3, len(annotations) + 3)}"


def _legacy_table_title(table: dict, chapter) -> str:
    for key in ("title", "caption"):
        value = table.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()

    for key in ("title", "caption"):
        value = table.get(key)
        text = _coerce_table_metadata_text(value)
        if text:
            return text

    chapter_title = _coerce_table_metadata_text(getattr(chapter, "title", "章节"))
    return f"{chapter_title or '章节'} 表格"


def _coerce_table_metadata_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (list, dict)):
        if not value:
            return ""
        try:
            return json.dumps(value, ensure_ascii=False).strip()
        except (TypeError, ValueError):
            return str(value).strip()
    return str(value).strip()


def _load_tables(content_json: str | None) -> list[dict]:
    if not content_json:
        return []
    try:
        data = json.loads(content_json)
    except (TypeError, json.JSONDecodeError):
        return []
    if not isinstance(data, dict):
        return []

    legacy_tables = data.get("tables", [])
    if isinstance(legacy_tables, list):
        valid_legacy_tables = [
            table
            for table in legacy_tables
            if (
                isinstance(table, dict)
                and isinstance(table.get("headers"), list)
                and table.get("headers")
                and isinstance(table.get("rows", []), list)
            )
        ]
        if valid_legacy_tables:
            return valid_legacy_tables

    content = data.get("content", [])
    if not isinstance(content, list):
        return []

    tables = []
    for index, node in enumerate(content):
        if not isinstance(node, dict) or node.get("type") != "table":
            continue
        table = _prosemirror_table_to_dict(node, _prosemirror_table_title(content, index))
        if table["headers"]:
            tables.append(table)
    return tables


def _prosemirror_table_to_dict(table_node: dict, title: str | None) -> dict:
    headers: list = []
    rows: list[list] = []
    header_found = False

    table_content = table_node.get("content", [])
    if not isinstance(table_content, list):
        return {"title": title or "", "headers": headers, "rows": rows}

    for row_node in table_content:
        if not isinstance(row_node, dict) or row_node.get("type") != "tableRow":
            continue
        row_content = row_node.get("content", [])
        if not isinstance(row_content, list):
            continue
        cell_nodes = [
            cell
            for cell in row_content
            if (
                isinstance(cell, dict)
                and cell.get("type") in {"tableCell", "tableHeader"}
                and isinstance(cell.get("content", []), list)
            )
        ]
        if not cell_nodes:
            continue

        header_cells = [cell for cell in cell_nodes if cell.get("type") == "tableHeader"]
        if header_cells and not header_found:
            headers = [_prosemirror_cell_value(cell) for cell in header_cells]
            header_found = True
            continue

        rows.append(
            [
                _prosemirror_cell_value(cell)
                for cell in cell_nodes
                if cell.get("type") in {"tableCell", "tableHeader"}
            ]
        )

    if not headers and rows:
        headers, rows = rows[0], rows[1:]

    return {"title": title or "", "headers": headers, "rows": rows}


def _prosemirror_table_title(content: list, table_index: int) -> str | None:
    table_node = content[table_index]
    attrs = table_node.get("attrs", {}) if isinstance(table_node, dict) else {}
    if isinstance(attrs, dict):
        for key in ("title", "caption"):
            value = attrs.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()

    if table_index == 0:
        return None
    previous = content[table_index - 1]
    if isinstance(previous, dict) and previous.get("type") == "paragraph":
        text = _prosemirror_node_text(previous).strip()
        return text or None
    return None


def _prosemirror_cell_value(cell_node: dict) -> str:
    paragraphs = []
    cell_content = cell_node.get("content", [])
    if not isinstance(cell_content, list):
        return ""

    for child in cell_content:
        if not isinstance(child, dict):
            continue
        text = _prosemirror_node_text(child)
        if child.get("type") == "paragraph" or text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def _prosemirror_node_text(node: dict) -> str:
    if not isinstance(node, dict):
        return ""
    if node.get("type") == "text":
        return str(node.get("text", ""))
    if node.get("type") == "hardBreak":
        return "\n"
    content = node.get("content", [])
    if not isinstance(content, list):
        return ""
    return "".join(
        _prosemirror_node_text(child)
        for child in content
        if isinstance(child, dict)
    )


def _normalize_row(row: Iterable) -> list:
    if not isinstance(row, Iterable) or isinstance(row, (str, bytes, dict)):
        return [_normalize_cell_value(row)]
    return [_normalize_cell_value(value) for value in row]


def _normalize_cell_value(value):
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, (datetime, date)):
        return _normalize_datetime(value) if isinstance(value, datetime) else value
    if isinstance(value, str):
        parsed = _parse_iso_date(value)
        if parsed is not None:
            return parsed
        stripped = value.strip()
        if stripped.lower() in {"true", "false"}:
            return stripped.lower() == "true"
        if re.fullmatch(r"[+-]?\d+", stripped):
            return int(stripped)
        if re.fullmatch(r"[+-]?(?:(?:\d+\.\d*|\.\d+)(?:[eE][+-]?\d+)?|\d+[eE][+-]?\d+)", stripped):
            return float(stripped)
        return value
    if isinstance(value, (list, tuple, set, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _parse_iso_date(value: str):
    try:
        if len(value) == 10 and value[4] == "-" and value[7] == "-":
            return date.fromisoformat(value)
        if "T" in value or " " in value:
            return _normalize_datetime(datetime.fromisoformat(value.replace("Z", "+00:00")))
    except ValueError:
        return None
    return None


def _normalize_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def _pad_row(row: list, size: int) -> list:
    return row + [None] * max(0, size - len(row))


def _sanitize_sheet_name(name: str) -> str:
    cleaned = "".join(ch for ch in name if ch not in "\\/:*?[]")
    cleaned = cleaned.strip() or "Sheet"
    return cleaned[:31]


def _excel_sheet_target(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''").replace('"', '""')
    return f"#'{escaped}'!A1"


def _unique_sheet_name(base_name: str, used_names: set[str]) -> str:
    candidate = base_name[:31] or "Sheet"
    used_names_casefolded = {name.casefold() for name in used_names}
    if candidate.casefold() not in used_names_casefolded:
        used_names.add(candidate)
        return candidate

    suffix_number = 2
    while True:
        suffix = f"-{suffix_number}"
        trimmed = base_name[: max(1, 31 - len(suffix))].rstrip("-")
        candidate = f"{trimmed}{suffix}"[:31]
        if candidate.casefold() not in used_names_casefolded:
            used_names.add(candidate)
            return candidate
        suffix_number += 1


def _set_column_widths(ws, max_cols: int, headers: list, rows: list[list]) -> None:
    all_rows = [headers, *rows]
    for col_index in range(1, max_cols + 1):
        values = []
        for row in all_rows:
            if col_index - 1 < len(row):
                values.append(row[col_index - 1])
        width = max([len(str(v)) for v in values if v not in (None, "")] or [10])
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(width + 2, 10), 28)


def _summarize_items(value) -> str:
    if not value:
        return "无"
    if isinstance(value, (str, bytes, dict)) or not isinstance(value, Iterable):
        items = [value]
    else:
        items = value

    normalized_items = []
    for item in items:
        if isinstance(item, dict):
            item = (
                item.get("description")
                or item.get("title")
                or item.get("name")
                or item
            )
        text = _normalize_metadata_text(item)
        if text:
            normalized_items.append(text)

    compact = [
        item if len(item) <= 60 else item[:57] + "..."
        for item in normalized_items[:5]
    ]
    if len(normalized_items) > 5:
        compact.append(f"等{len(normalized_items) - 5}项")
    return "；".join(compact) or "无"


def _normalize_metadata_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (list, tuple, set, dict)):
        try:
            return json.dumps(value, ensure_ascii=False, default=str).strip()
        except (TypeError, ValueError, RecursionError):
            pass
    return str(value).strip()


def _active_export_annotations(annotations: list | None) -> list[dict]:
    normalized = []
    for annotation in annotations or []:
        if _annotation_value(annotation, "status", "pending") == "ignored":
            continue
        target_text = str(_annotation_value(annotation, "target_text", "") or "").strip()
        content = str(_annotation_value(annotation, "content", "") or "").strip()
        if not target_text and not content:
            continue
        normalized.append(
            {
                "id": _annotation_value(annotation, "id", ""),
                "chapter_id": _annotation_value(annotation, "chapter_id", ""),
                "chapter_title": _annotation_value(annotation, "chapter_title", ""),
                "target_text": target_text,
                "content": content or "（未填写批示内容）",
                "status": _annotation_value(annotation, "status", "pending") or "pending",
                "created_by": _annotation_value(annotation, "created_by", "本地用户") or "本地用户",
                "locator": _annotation_value(annotation, "locator", "") or "",
            }
        )
    return normalized


def _annotations_by_chapter(annotations: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = {}
    for annotation in annotations:
        chapter_id = str(annotation.get("chapter_id") or "")
        grouped.setdefault(chapter_id, []).append(annotation)
    return grouped


def _annotation_value(annotation, name: str, default=None):
    if isinstance(annotation, dict):
        return annotation.get(name, default)
    return getattr(annotation, name, default)
